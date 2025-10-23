# app.py
from flask import Flask, render_template, request
from datetime import datetime
from logic_ai import analyze_survey_results, HOLLAND_MAPPING
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os

app = Flask(__name__)

# Tên file để lưu trữ kết quả khảo sát
DATA_FILE = 'data.xlsx'

@app.route('/')
def index():
    """Trang chủ của ứng dụng."""
    return render_template('index.html')

@app.route('/khaosat')
def survey():
    """Trang hiển thị form khảo sát."""
    return render_template('khaosat.html', holland_questions=HOLLAND_MAPPING)

@app.route('/ketqua', methods=['POST'])
def result():
    """Trang xử lý dữ liệu, lưu kết quả và hiển thị."""
    try:
        # 1. Thu thập dữ liệu thông tin cá nhân
        personal_info = {
            "ho_ten": request.form.get("ho_ten"),
            "lop": request.form.get("lop"),
            "sdt": request.form.get("sdt"),
            "email": request.form.get("email"),
            "thoi_gian": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        # 2. Thu thập dữ liệu khảo sát
        survey_data = {
            "part_a": {},
            "part_b": {
                "cau2": request.form.getlist("cau2"),
                "cau3": request.form.getlist("cau3"),
                "cau4": [request.form.get("cau4")]
            },
            "part_c": {"cau5": request.form.getlist("cau5")}
        }
        for key in request.form:
            if key.startswith("mon_"):
                subject_name = key.replace("mon_", "")
                survey_data["part_a"][subject_name] = int(request.form[key])

        # 3. Gọi hàm phân tích từ logic_ai.py
        recommendations = analyze_survey_results(survey_data)
        
        # 4. Lưu tất cả thông tin vào file data.xlsx
        save_results_to_excel(personal_info, recommendations)
        
        # 5. Trả về trang kết quả
        return render_template('ketqua.html', recommendation_text=recommendations)

    except Exception as e:
        return f"Đã xảy ra lỗi, có thể bạn đã bỏ sót một câu hỏi nào đó. Vui lòng quay lại và thử lại. Lỗi: {e}"

def save_results_to_excel(personal_info, recommendations):
    """Hàm để lưu kết quả khảo sát vào file Excel."""
    
    # Kiểm tra xem file đã tồn tại chưa
    if not os.path.exists(DATA_FILE):
        # Nếu chưa, tạo file mới và ghi hàng tiêu đề
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "KetQuaKhaoSat"
        header = ["Timestamp", "Họ và Tên", "Lớp", "SĐT", "Email", "Gợi ý 1", "Gợi ý 2", "Gợi ý 3"]
        sheet.append(header)
        
        # Tự động điều chỉnh độ rộng cột cho đẹp
        for i, column_title in enumerate(header, 1):
            sheet.column_dimensions[get_column_letter(i)].best_fit = True
    else:
        # Nếu file đã tồn tại, mở nó lên
        workbook = openpyxl.load_workbook(DATA_FILE)
        sheet = workbook.active

    # Chuẩn bị dữ liệu hàng mới
    rec_names = [rec.get('name', '') for rec in recommendations.get('recommendations', [])]
    rec_names += [''] * (3 - len(rec_names)) # Đảm bảo luôn có 3 phần tử

    row_data = [
        personal_info["thoi_gian"],
        personal_info["ho_ten"],
        personal_info["lop"],
        personal_info["sdt"],
        personal_info["email"],
        rec_names[0],
        rec_names[1],
        rec_names[2]
    ]
    
    # Thêm hàng dữ liệu mới vào file
    sheet.append(row_data)

    # Lưu lại file
    workbook.save(DATA_FILE)

from flask import Flask
import os

app = Flask(__name__)

@app.route('/')
def home():
    return "Flask app is running on Render!"

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)

